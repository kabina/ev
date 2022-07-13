#!/usr/bin/python
# -*- coding: utf-8 -*-
"""전기차 충전기 시뮬레이터

     * 안창선
     * 2022-05-27

Todo:
   TODO리스트를 기재
    * 비정상 케이스에 대한 프로세스 적용
    * ID값들 자동화 처리 필요


"""
import random
from abc import *
from datetime import datetime
import requests
import json
import props
import scenario
import time, os
import urllib3
import logging
from evlogger import Logger
import multiprocessing
from tqdm import tqdm


"""랜덤으로 ID태깅이 이루어 지는 경우 사용될 테스트 ID들
"""
idTags = None
SLEEP = 1
logger = Logger()
global evlogger
evlogger = logger.initLogger(loglevel=logging.INFO)
conn = None

class Server(metaclass=ABCMeta):
    @abstractmethod
    def make_request(self, request_type):
        pass


class Charger(Server):
    """충전기 클래스 선언

     충전기는 CS로 요청을 보내기도 하고, 자체 서버를 띄워 요청을 받을 수도 있어야 함(HTTP)

    Attributes:
        Server (Class): 서버 클래스의 abstract class inherit

    """
    def __init__(self, charger_id=None):

        if charger_id is None or len(charger_id) != 13:
            evlogger.error("충전기ID 오류로 충전기를 생성 할 수 없습니다.")
            return

        self.stop_by_error = False  # CS로 부터 Response 정보가 부족하더라도 시뮬레이트 계속 작동
        self.local_var = {
            "transactionId": None,
            "heartbeatInterval": 5,
            "vendorId": "LGE",
            "connectorId": None,
            "status": "Available",
            "errorCode": "NoError",
            "bootNotification_reason": None,
            "stopTransaction_reason": None,
            "statusNotification_reason": None,
            "cmeter":0,
            "meterStart": 0,
            "meterStop": 0,
            "idTag": None,
            "X-EVC-MDL": "LGE-123",
            "X-EVC-BOX": None,
            "X-EVC-OS": "Linux 5.5",
            "tariff": [],
            "responseFailure": False,
        }

        # meterValue용 파라메터 변수 들 (기본값)

        self.sampled_value = {
            "cimport": 15,  # Current.Import, 충전전류(A)
            "voltage": 220.0,  # Voltage
            "eairegister": 0,  # Energy.Active.Import.Register (Wh)
            "soc": 10,  # SoC
            "paimport": 0,  # Power.Active.Import 충전기로 지속 충전되는 양 (W)
        }

        self.local_var["X-EVC-BOX"] = charger_id[:13]
        self.local_var["connectorId"] = charger_id[11:12]
        self.local_var["idTag"] = random.choice(idTags)

        self.setter = {
            "timestamp": lambda x: datetime.now().replace(microsecond=0).isoformat(),
            "transactionId": lambda x: self.local_var[x],
            "vendorId": lambda x: self.local_var[x],
            "connectorId": lambda x: self.local_var[x],
            "status": lambda x: self.local_var[x],
            "meterStart": lambda x: self.local_var[x],
            "meterStop": lambda x: self.local_var[x],
            "errorCode": lambda x: self.local_var[x],
            "bootNotification_reason": lambda x: self.local_var[x],
            "idTag": lambda x: self.local_var[x],
            "stopTransaction_reason": lambda x: self.local_var[x],
            "statusNotification_reason": lambda x: self.local_var[x],
        }

    def reset_charger(self, charger_id, idTags):
        self.local_var["X-EVC-BOX"] = charger_id[:13]
        self.local_var["connectorId"] = charger_id[11:12]
        self.local_var["idTag"] = random.choice(idTags)

    def init_local_var(self, ):
        fixed = [
            "vendorId",
            "X-EVC-BOX",
            "X-EVC-MDL",
            "X-EVC-OS",
            "heartbeatInterval",
            "connectorId",
            "cmeter",
            "meterStart",
            "meterStop",
            "idTag"
        ]
        for item in self.local_var:
            if item not in fixed:
                self.local_var[item] = None

    def disp_header(self, command):
        evlogger.info("#" * 50)
        evlogger.info("############# - Request for {}".format(command))
        evlogger.info("#" * 50)

    def update_var(self, item, value):
        self.local_var[item] = value

    def login(self, userid, password):

        data = {"userId": userid, "userPw": password}
        response = requests.post(props.urls["login"], headers=props.headers,
                                 data=json.dumps(data))
        response_dict = json.loads(response.text)
        self.accessToken = response_dict["payload"]["accessToken"]

    def meter_update(self, command=None):
        self.sampled_value["cimport"] += random.uniform(-1, 1)
        self.sampled_value["voltage"] = self.sampled_value["voltage"] + random.uniform(-1, 1)
        self.sampled_value["eairegister"] = (self.sampled_value["eairegister"] +
                                             random.randrange(9950, 10500)) if command == "meterValues" else 0
        self.sampled_value["soc"] = self.sampled_value["soc"]
        self.sampled_value["paimport"] += random.uniform(-1, 1)

        self.update_var("cmeter", self.local_var["meterStart"]+self.sampled_value["eairegister"])

    def get_sampledValue(self):
        self.meter_update(command="meterValues")
        return [
            {"measurand": "Current.Import", "phase": "L1", "unit": "A", "value":
                "{:.1f}".format(self.sampled_value["cimport"]),  # Wh Value
             }, {"measurand": "Voltage", "phase": "L1", "unit": "V", "value":
                "{:.1f}".format(self.sampled_value["voltage"]),
                 }, {"measurand": "Energy.Active.Import.Register", "unit": "Wh", "value":
                "{:.1f}".format(self.sampled_value["eairegister"])
                     }, {"measurand": "SoC", "unit": "%", "value":
                "{}".format(self.sampled_value["soc"], )
                         }, {"measurand": "Power.Active.Import", "unit": "W", "value":
                "{:.1f}".format(self.sampled_value["paimport"])
                             }
        ]

    def send_response(response):
        print(response)

    def ltouch_parameter(self, parameter, command=None):
        """Request시 정의된 JSON내 변경이 필요한 항목을 탐색하여 수정함.
        timestamp나 transactionId등을 해당 시점에 맞는 값으로 변경 해 줌

        Args:
          parameter: CS로 송신되는 파라메터값, Recursive하게 들어옴.
          command: 충전기->CS, 요청(boot, statusNotification, startTransaction, ... )

        Returns:
          변경된 Parameter.

        Raises:
          None.
        """
        new_param = []
        for item in parameter:
            if type(item) is dict:
                new_param.append(self.touch_parameter(item, command=command))
            elif type(item) is list:
                new_param.append(self.ltouch_parameter(item, command=command))
            elif item in self.setter:
                new_param.append(self.setter[item](item))
            else:
                new_param.append(item)
        return new_param

    def touch_parameter(self, parameter, command=None):
        """Request시 정의된 JSON내 변경이 필요한 항목을 탐색하여 수정함.
        timestamp나 transactionId등을 해당 시점에 맞는 값으로 변경 해 줌

        Args:
          parameter: CS로 송신되는 파라메터값, Recursive하게 들어옴.
          command: 충전기에서 CS로 보내는 명령 종류(bootNotification, Authorize, ... )

        Returns:
          변경된 Parameter.

        Raises:
          None.
        """
        new_param = {}
        for item in parameter.items():
            if type(item[1]) is dict:
                new_param[item[0]] = self.touch_parameter(item[1], command=command)
            elif type(item[1]) is list:
                if command == "meterValues" and item[0] == "sampledValue":
                    new_param[item[0]] = self.get_sampledValue()
                else:
                    new_param[item[0]] = self.ltouch_parameter(item[1], command=command)
            elif item[0] == "reason" and command in ["bootNotification", "stopTransaction", "statusNotification"]:
                new_param[item[0]] = self.setter[command + "_" + item[0]](command + "_" + item[0])
            elif item[0] in self.setter:
                new_param[item[0]] = self.setter[item[0]](item[0])
            else:
                new_param[item[0]] = item[1]
        return new_param

    def check_response(self, diff_from_items, diff_to_items):
        # from_item : response
        # to_item : props
        """Response 데이터의 포맷 및 필수값 확인, Response값 중 충전기 내부적으로
        관리 해야 하는 데이터는 로컬 변수에 저장(트랜젹선ID 등).

        Args:
          diff_from_items: 실제 충전기로 전달 된 응답값
          diff_to_items: 응답값과 비교 해야 할 표준 응답 포맷 및 필수여부, 저장 해야 할 키 값

        Returns:
          None.

        Raises:
          None.
        """
        for to_item in diff_to_items.items():
            # 데이터타입이 Dict인 경우
            if type(to_item[1]) is dict:
                if to_item[0] in diff_from_items:
                    self.check_response(diff_from_items[to_item[0]], to_item[1])
                else:
                    msg = "No Response Item '{}'".format(to_item[0])
                    evlogger.error(msg)
                    if self.stop_by_error:
                        raise Exception(msg)
            # 데이터타입이 List인 경우 (반드시 Element는 1개 이상의 Dict임)
            elif type(to_item[1]) is list and type(to_item[1][0]) is dict:
                # tariff와 같이 하위 List Element내에 dict가 반복되는 데이터셋 처리
                if to_item[0] in diff_from_items:
                    for idx in range(len(to_item)):
                        if len(diff_from_items[to_item[0]]) > 0 and diff_from_items[to_item[0]][idx] is dict:
                            self.check_response(diff_from_items[to_item[0]][idx], to_item[1][idx])
                        else:
                            break
                # evlogger.info("CS Not ready for list in response")
            elif to_item[0] in diff_from_items:
                if to_item[1][1]:
                    self.update_var(to_item[1][1], diff_from_items[to_item[0]])
                elif to_item[0] in self.setter:
                    self.local_var[to_item[0]] = diff_from_items[to_item[0]]
            else:
                if to_item[1][0] == "M":
                    evlogger.error("No Response Item '{}'".format(to_item[0]))
                    self.update_var("responseFailure", True)
                else:
                    evlogger.warning("No Response Item '{}'(Optional)".format(to_item[0]))

    def resp_post_process(self, response, command):
        """충전기는 CS(Central System)로 부터 받은 응답을 후처리 함

        Args:
          response : CS로 부터 받은 응답
          command : 충전기에서 서버로 보낸 명령(authorize, boot, startTransactionRequest ... )

        Returns:
          errorCode: -1 : No error
                      > 1 : Error
          errorMsg: error message.

        Raises:
          None.cmd
        """
        error_code = -1
        error_msg = None

        # 사용자 요금정보(Tariff) 처리
        if command == "dataTransferTariff":
            if "data" in response and "tariff" in response["data"] :
                self.local_var["tariff"] = response["data"]["tariff"]
        if command == "startTransaction":
            if response["transactionId"] is None or len(response["transactionId"]) == 0:
                evlogger.error("No Transaction ID")
                error_msg = "No Transaction ID"
                error_code = 1
            else:
                self.local_var["transactionId"] = response["transactionId"]
        if command in ["authorize"] and "idTagInfo" in response:
            self.local_var["status"] = response["idTagInfo"]["status"]

        if command == "stopTransaction" and "idTagInfo" in response:
            self.local_var["status"] = response["idTagInfo"]["status"]

        if command == "bootNotification" and "status" in response:
            self.local_var["status"] = response["status"]


        if error_code > 0:
            raise Exception(error_msg)

    def req_post_process(self, requests, command):
        """충전기는 CS(Central System)로 부터 보내는 요청(Requests) 생성 후 local variable 등 Update 처리

        Args:
          requests : CS로 보내는 요청
          command : 충전기에서 서버로 보낸 명령(authorize, boot, startTransactionRequest ... )

        Returns:
          errorCode: -1 : No error
                      > 1 : Error
          errorMsg: error message.

        Raises:
          None.cmd
        """

        if command == "statusNotification" and "status" in requests:
            self.local_var["status"] = requests["status"]
        if command == "authorize" and "idTag" in requests:
            self.local_var["idTag"] = requests["idTag"]
            self.sampled_value["eairegister"] = 0
            self.local_var["meterStart"] = self.local_var["cmeter"]
        if command == "stopTransaction" :

            self.local_var["meterStop"] = self.local_var["cmeter"]
            requests["meterStop"] = self.local_var["meterStop"]


    def make_request(self, command=None, status=None):
        """충전기에서 CS(Central System)으로 보낼 데이터를 생성 하고 송신 후 Response를 수신 함
        수신된 Response는 포맷/필수여부/로컬변수 저장등을 위해 check_response함수 호출

        Args:
          command : 충전기에서 서버로 보낼 명령(authorize, boot, startTransactionRequest ... )
          status : 명령 수행 후 변경 할 상태 지정(Available, Charging, Preparing, ... )

        Returns:
          None.

        Raises:
          None.
        """

        url = props.urls[command]
        parameter = props.api_params[command]

        def set_header(header, headers):
            if command == "bootNotification":
                ri = "boot"
            elif command == "dataTransferHeartbeat":
                ri = "heartbeat"
            else:
                ri = "card"
            get_headers = {
                'X-EVC-RI': datetime.now().strftime('%Y%m%d%H%M%S%f')[:-3] + "_" + ri,
                'X-EVC-MDL': self.local_var["X-EVC-MDL"],
                'X-EVC-BOX': self.local_var["X-EVC-BOX"],
                'X-EVC-OS': self.local_var["X-EVC-OS"],
                'X-EVC-CON': self.local_var["connectorId"],
            }
            for h in headers:
                header[h] = get_headers[h]
            return header

        # 기본 템플릿 기반으로 파라미터 설정
        # if command == "stopTransaction" :
        #     self.local_var["idTag"] = "4634407056130185"
        parameter = self.touch_parameter(parameter, command=command)


        # if status is not None:
        #     parameter["status"] = status

        self.disp_header(command)
        header = props.headers

        self.req_post_process(parameter, command)

        # header["Authorization"] = "Bearer {}".format(self.accessToken)
        if command in ["bootNotification", "authorize", "dataTransferHeartbeat", "statusNotification"]:
            header = set_header(header, props.api_headers[command])

        data = json.dumps(parameter)

        evlogger.info(">" * 10 + "송신 헤더" + ">" * 10)
        evlogger.info(header)
        evlogger.info(">" * 10 + "송신 BODY" + ">" * 10)
        evlogger.info(data)

        response = requests.post(url, headers=header, data=data, verify=False)

        if response.status_code in [503, 404, 403, 500, 401]:
            evlogger.error("Internal Service Error or No Available Service. [{}]".format(response.status_code))
            self.local_var["status"] = "ServerError"
        else:
            response = response.json()

            evlogger.info("<" * 10 + "수신 DATA" + "<" * 10)
            evlogger.info(response)

            if command not in props.api_response:
                evlogger.warning("NO response checker for {}".format(command))
            else:
                self.check_response(response, props.api_response[command])

            # 정상 Response를 받은 경우 아래 수행
            if not self.local_var["responseFailure"]:
                self.resp_post_process(response, command)

        return [self.local_var["X-EVC-BOX"],
                self.local_var["idTag"],
                header["X-EVC-RI"],
                self.local_var["status"],
                command,
                self.local_var["transactionId"],
                self.local_var["meterStart"],
                self.local_var["meterStop"],
                self.local_var["cmeter"],
                (max(self.local_var["meterStop"],self.local_var["cmeter"]) - self.local_var["meterStart"]) if command == "stopTransaction" else 0,
                str(parameter),
                str(response),
                str(header),
                ]


def case_run(charger, case) -> list:
    """충전기 기본 시뮬레이트 실행. 시험 케이스에 따라 충전기 동작 수행
    Args:
      charger : 충전기 instance

    Returns:
      out_list : request별 transaction관련 in/out 내역.

    Raises:
      None.
    """

    # param 1: 충전기 인스턴스
    # param 2: 시험 케이스
    # 케이스 별로 사전 상태변수 및 오류코드 세팅 후 Request 요청

    charger.init_local_var()

    out_list = []
    pbar = tqdm(case)
    metercount = case.count(['meterValues'])
    mc = 1
    for task in pbar:
        if task[0] == 'meterValues':
            pbar.set_description_str(f'Processing : {"meterValues":17} {mc}/{metercount}')
            mc += 1
        else:
            pbar.set_description_str(f'Processing : {task[0]:21}')

        if task[0] == "statusNotification":
            if len(task) > 2:  # 2nd element(arg)가 있는 경우만
                charger.update_var("errorCode", task[2])
                charger.update_var("statusNotification_reason", task[3])
            charger.update_var("status", task[1])
        elif task[0] in ["bootNotification", "stopTransaction"]:
            # 부팅, 종료(정지) 이유 등록
            charger.update_var(task[0] + "_reason", task[1])
        if task[0] == "authorize":
            charger.local_var["meterStart"] = charger.local_var["cmeter"]

        if task[0] == "meterValues":
            for _ in range(random.randrange(1,10)):
                row = charger.make_request(command=task[0])
                time.sleep(SLEEP)
                out_list.append(row)
        else:
            row = charger.make_request(command=task[0])
            out_list.append(row)

        time.sleep(SLEEP)
        evlogger.info("=" * 20 + "최종 충전기 내부 변수 상태" + "=" * 18)
        evlogger.info(charger.local_var)
        evlogger.info("=" * 60)
        # time.sleep(0.1)

    return out_list

def main(charger_id="charger_01"):
    # exclude SSL Warning message
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection

    # evlogger = logger.initLogger(logid=charger_id)

    wb = Workbook()  # create xlsx file
    ws = wb.active

    ws.append(["Charger_ID", "Card_Id", "RI", "Status", "Command",
               "Transaction_ID", "MeterStart", "MeterStop", "Meter", "Charge_Amount", "Request Json", "response Json",
               "Header"])
    """ 9자리 : 충전소 (115000001)
        2자리 : 충전기 (01)
        1자리 : 커넥터 (0)
        1자리 : 충전타입(급/중/저) (A/B/C)
    """
    charger = Charger(charger_id=crgrList[random.randrange(0,len(crgrList))])
    # charger = Charger("010400001", "010400001100A", "01")

    """초기 충전기 미터값 설정
    """
    meterstart = random.randrange(100_000, 120_000)
    charger.update_var("meterStart",meterstart)
    charger.update_var("cmeter", meterstart)
    # charger.update_var("cmeter", charger.local_var["meterStop"])

    loop_cnt = 1
    print(f"시나리오: 총 {loop_cnt}회 충전 수행")

    # for l in case_run(charger, scenario.coupler_connect):
    #     ws.append(l)

    for _ in range(loop_cnt):
        for l in case_run(charger, scenario.normal_case):
            ws.append(l)

        # for l in case_run(charger, scenario.normal_case_reserved):
        #         ws.append(l)
        #
        # for l in case_run(charger, scenario.invalid_card_in_reserved):
        #         ws.append(l)
        #
        # for l in case_run(charger, scenario.normal_case_ing):
        #         ws.append(l)
        #
        # for l in case_run(charger, scenario.normal_case_without_boot):
        #     ws.append(l)
        #
        # for l in case_run(charger, scenario.error_in_charge):
        #     ws.append(l)
        charger.reset_charger(crgrList[random.randrange(0,len(crgrList))], idTags)
    # ws.append(case_run(charger, scenario.error_just_after_boot))
    # ws.append(case_run(charger, scenario.heartbeat_after_boot))
    # ws.append(case_run(charger, scenario.no_charge_after_authorize))
    # case_run(scenario.reserved_after_boot)
    # case_run(scenario.remote_stop_transaction)

    wsum = charger.local_var["meterStop"] - meterstart
    print(f"총 {wsum:,} Wh의 전력 충전")

    """Excel Cell width and style setting
    """
    cell_width = [15, 20, 25, 10, 15, 18, 10, 10, 10, 10, 40, 40, 40]
    for idx, w in enumerate(cell_width):
        ws.column_dimensions[chr(ord("A") + idx)].width = w
    border = Border(bottom=Side(style="thick"))
    for cell in ws["1:1"]:
        cell.border = border

    """엑셀저장 할 파일명 지정, 저장 전에 기존 파일 존재 할 경우, 삭제 후 저장 함
    """
    xlsx_file = charger_id+".xlsx"
    if os.path.isfile(xlsx_file):
        os.remove(xlsx_file)
    wb.save(xlsx_file)

# 작업 리스트를 반환
def getWorkList():
    work_list = []

    for i in tqdm(range(0, 1)):
        work_list.append('charger_' + str(i))

    return work_list

def getConnection():
    import pymysql
    conn = pymysql.connect(host="rds-aurora-mysql-ev-charger-svc-instance-0.cnjsh2ose5fj.ap-northeast-2.rds.amazonaws.com",
                           user='evsp_usr', password='evspuser!!', db='evsp', charset='utf8', port=3306)

    return conn


def getCards():

    with conn.cursor() as cur:
        cur.execute(" select b.mbr_card_no "+
                    " from mbr_info a "+
                    " inner join mbr_card_isu_info b "+
                    " on a.mbr_id = b.mbr_id "+
                    " where b.card_stus_cd = '01'")
        return cur.fetchall()


def getCrgrs(chrstn_id = None):

    with conn.cursor() as cur:
        sql = " select b.crgr_cid " \
              " from crgr_mstr_info a " \
              " inner join crgr_info b " \
              " on a.crgr_mid = b.crgr_mid " \
              " where a.crgr_stus_cd = '04' and b.crgr_cid like '%A'"

        if chrstn_id :
            sql = sql + f" and a.chrstn_id = '{chrstn_id}' "
        cur.execute(sql)

        return cur.fetchall()

def getMCrgrs(chrstn_id = None):

    with conn.cursor() as cur:
        sql = " select crgr_mid " \
              " from crgr_mstr_info "

        if chrstn_id :
            sql = sql + f" where chrstn_id = '{chrstn_id}' "
        cur.execute(sql)

        return cur.fetchall()

if __name__ == "__main__":
    DBCONN = False

    if DBCONN :
        conn = getConnection()
        idTags = random.sample([card[0] for card in getCards()], k=100)
        crgrList = random.sample([crgr[0] for crgr in getCrgrs()], k=100)
        conn.close()
    else:
        """
        5555222233334444, 3333222233334444(정지카드), 1010202030306060, 1010010174366716(미가입카드), 1010010174721340(가입카드)
        9999999999999999 비회원(APP Case)
        4873600231574325 (cust012 - 고정요금)
        4748664213640739 (cust013 - 고정-멤버십(일반))
        4200293606526754 (cust014 - 고정-멤버십(VIP))
        4634407056130185 (cust015 - 구독요금제)
        1010010174366716 (로밍-미가입카드)
        1010010174721340 (로밍-가입카드)
        4677893456241782 (안창선)
        4677893456241733 (안창선New)
        crgr_mid;crgr_cid;me_crgr_id;chrstn_id;chrstn_cntc_no;crgr_open_yn;prcl_mttr;regt_id;reg_dttm;mfpn_id;upd_dttm
        98001010101;980010101010A;01;112000222;
        98001010101;980010101020A;02;112000222;
        98001010101;980010101030A;03;112000222;
        98001010201;980010102010A;01;112000222;
        98001010202;980010102020A;02;112000222;
        98001010203;980010102030A;03;112000222;
        98001020101;980010201010A;01;112000222;
        98001020102;980010201020A;02;112000222;
        98001030101;980010301010A;01;112000222;
        """
        idTags = ['4677893456241733', #'3333222233334444', '1010010174721340',
                  #'1010010174366716', # 환경부 중지 카드
                  #'1010010107409567', # 환경부 정상 카드
                  #'1010010177777471', # 환경부 정상 카드(협약사)
                  #'4873600231574325',
                  #'4748664213640739', '4634407056130185'
                  ] # 1111222233334444, 4873600231574325(cust012-고정요금)
        crgrList = ['115000001010A' #, '114100005090A', '112000006240A'
                    ] # 112000006240A, 114100005030A, 115000001010A, 112000006220B, 112000012010A, 114100005090A(예약)

    main()

    # try :
    #     pool = multiprocessing.Pool(processes=1)  # 3개의 processes 사용
    #     pool.map(main, getWorkList())
    #     pool.close()
    #     pool.join()
    # except PermissionError as e:
    #     print("output.xlsx 파일이 사용 중입니다. 쓰기 실패.")
    #
